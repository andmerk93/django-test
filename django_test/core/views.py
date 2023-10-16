from io import BytesIO
from datetime import datetime as dt
from itertools import islice

from django.http import StreamingHttpResponse
from django.shortcuts import get_object_or_404, get_list_or_404
from django.conf import settings
from django.utils import timezone

from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import (
    api_view, permission_classes, parser_classes
)
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework import status

from pyowm import OWM
from openpyxl import load_workbook
import xlsxwriter

from .models import Post, Place
from .serializers import PostSerializer, PlaceSerializer, WeatherSerializer

MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class PostViewSet(ModelViewSet):
    queryset = Post.objects.all()
    serializer_class = PostSerializer


class PlaceViewSet(ModelViewSet):
    queryset = Place.objects.all()
    serializer_class = PlaceSerializer


def import_weather(place_name, author_id):
    place = get_object_or_404(Place, title=place_name)
    weather = OWM(settings.OWM_API_KEY).weather_manager().weather_at_coords(
        lat=float(place.point.latitude),
        lon=float(place.point.longitude),
    ).weather
    # longitude latitude
    hhHg_pressure = weather.barometric_pressure().get('press') * 0.750064  # noqa
    # 1 hPa = 0.750064 hhHg
    data = dict(
        temper=weather.temperature('celsius').get('temp'),
        humidity=weather.humidity,
        pressure=int(hhHg_pressure),
        wind_direction=weather.wind().get('deg'),  # degrees
        wind_speed=weather.wind().get('speed'),    # m / s
        place=place.id,
        author=author_id,
    )
    return WeatherSerializer(data=data)


def serializer_saver(serializer):
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(
        serializer.errors,
        status=status.HTTP_500_INTERNAL_SERVER_ERROR
    )


@api_view()
@permission_classes([IsAuthenticated])
def import_weather_manually(request, place_name):
    author_id = request.user.id
    serializer = import_weather(place_name, author_id)
    return serializer_saver(serializer)


@api_view()
@permission_classes([IsAuthenticated])
def export_weather(request, place_name):
    date_from_query = request.query_params.get('date')
    if not date_from_query:
        return Response(
            'Задайте дату в формате YYYY-MM-DD',
            status=status.HTTP_400_BAD_REQUEST
        )
    try:
        date_from_query = tuple(map(int, date_from_query.split('-')))
        date = dt(
            year=date_from_query[0],
            month=date_from_query[1],
            day=date_from_query[2],
            tzinfo=timezone.get_current_timezone(),
        ).date()
    except Exception:
        return Response(
            'Задайте дату в формате YYYY-MM-DD',
            status=status.HTTP_400_BAD_REQUEST
        )
    place = get_object_or_404(Place, title=place_name)
    weathers = get_list_or_404(place.weathers, date__date=date)
    serializer = WeatherSerializer(weathers, many=True)
    file_name = f'weather_at_{place_name}-{date}.xlsx'
    byte_data = BytesIO()
    with xlsxwriter.Workbook(byte_data, {'in_memory': True}) as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0, 0, serializer.data[0].keys())
        for num, data in enumerate(serializer.data, 1):
            worksheet.write_row(num, 0, data.values())
    byte_data.seek(0)
    return StreamingHttpResponse(
        streaming_content=byte_data,
        headers={
            'Content-Disposition': f'attachment; filename="{file_name}"',
        },
        content_type=MIME,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([FileUploadParser])
def import_places(request):
    file_obj = request.data['file']
    wb = load_workbook(filename=file_obj, read_only=True)
    excel_headers = next(wb.active.values)
    results = [
        dict(zip(excel_headers, row))
        for row in islice(wb.active.values, 1, None)
    ]
    serializer = PlaceSerializer(
        data=results,
        context=dict(request=request),
        many=True,
    )
    return serializer_saver(serializer)
